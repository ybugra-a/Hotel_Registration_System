"""
Veri Yonetimi - Excel okuma/yazma islemleri
"""

import os
import json
from datetime import datetime, date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Sabit yollar
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
CONFIG_DIR = os.path.join(BASE_DIR, "config")
EXCEL_PATH = os.path.join(DATA_DIR, "kayitlar.xlsx")
CONFIG_PATH = os.path.join(CONFIG_DIR, "ayarlar.cfg")

# Excel sutun indeksleri
COL_ID = 1       # A
COL_TC = 2       # B
COL_ISIM = 3     # C
COL_SOYISIM = 4  # D
COL_ODA = 5      # E
COL_GIRIS = 6    # F
COL_CIKIS = 7    # G
COL_ODEME_YON = 8  # H
COL_ODEME_TUT = 9  # I
COL_DURUM = 10     # J

ODEME_YONTEMLERI = ["Nakit", "Kredi Karti", "Banka Karti"]
DONEMLER = {1: "Q1", 2: "Q1", 3: "Q1",
            4: "Q2", 5: "Q2", 6: "Q2",
            7: "Q3", 8: "Q3", 9: "Q3",
            10: "Q4", 11: "Q4", 12: "Q4"}


class DataManager:
    def __init__(self):
        self._ensure_dirs()
        self._ensure_excel()
        self._ensure_config()

    def _ensure_dirs(self):
        os.makedirs(DATA_DIR, exist_ok=True)
        os.makedirs(CONFIG_DIR, exist_ok=True)

    def _ensure_excel(self):
        if not os.path.exists(EXCEL_PATH):
            self._create_new_excel()

    def _ensure_config(self):
        if not os.path.exists(CONFIG_PATH):
            self._save_config({"son_yedekleme": None, "version": "1.0"})

    def _create_new_excel(self):
        wb = Workbook()
        # Default sheti kaldir
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Ilk donem sheeti olustur
        now = datetime.now()
        sheet_name = f"{now.year}_{DONEMLER[now.month]}"
        ws = wb.create_sheet(sheet_name)
        self._create_sheet_headers(ws)

        # Odalar sheeti olustur
        ws_oda = wb.create_sheet("Odalar")
        ws_oda.column_dimensions['A'].width = 15
        ws_oda.column_dimensions['B'].width = 15
        ws_oda["A1"] = "Oda Numarasi"
        ws_oda["B1"] = "Durum"
        self._style_header_row(ws_oda, 1)

        # Ornek odalar ekle (101-110)
        for i in range(101, 111):
            ws_oda.append([str(i), "Musait"])

        wb.save(EXCEL_PATH)

    def _create_sheet_headers(self, ws):
        headers = [
            "Kayit ID", "T.C Kimlik No", "Isim", "Soyisim",
            "Oda Numarasi", "Giris Tarihi", "Cikis Tarihi",
            "Odeme Yontemi", "Odeme Tutari", "Durum"
        ]
        ws.append(headers)
        self._style_header_row(ws, 1)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14

    def _style_header_row(self, ws, row):
        header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _load_config(self):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"son_yedekleme": None, "version": "1.0"}

    def _save_config(self, config):
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

    def get_excel_path(self):
        return EXCEL_PATH

    def get_or_create_sheet(self, year, month):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet_name = f"{year}_{DONEMLER[month]}"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            self._create_sheet_headers(ws)
            wb.save(EXCEL_PATH)
        return sheet_name

    def get_next_id(self):
        """Tum sheetlerdeki en buyuk ID'yi bul ve +1 don"""
        wb = openpyxl.load_workbook(EXCEL_PATH)
        max_id = 0
        for sname in wb.sheetnames:
            if sname == "Odalar":
                continue
            ws = wb[sname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).isdigit():
                    max_id = max(max_id, int(row[0]))
        return max_id + 1

    def kayit_ekle(self, tc, isim, soyisim, oda, giris, cikis, odeme_yon, odeme_tut):
        """Yeni musteri kaydeder"""
        now = datetime.now()
        sheet_name = self.get_or_create_sheet(now.year, now.month)
        kayit_id = self.get_next_id()

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]

        giris_str = giris.strftime("%d/%m/%Y") if hasattr(giris, 'strftime') else str(giris)
        cikis_str = cikis.strftime("%d/%m/%Y") if (cikis and hasattr(cikis, 'strftime')) else (str(cikis) if cikis else "")
        odeme_str = str(odeme_tut).replace(',', '.') if odeme_tut else ""

        ws.append([
            kayit_id, tc, isim, soyisim, oda,
            giris_str, cikis_str, odeme_yon, odeme_str, "Aktif"
        ])

        # Oda durumunu guncelle
        self._update_oda_durumu(wb, oda, "Dolu")
        wb.save(EXCEL_PATH)
        return kayit_id

    def _update_oda_durumu(self, wb, oda_no, durum):
        if "Odalar" not in wb.sheetnames:
            return
        ws = wb["Odalar"]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(oda_no):
                row[1].value = durum
                return

    def get_aktif_musteriler(self):
        """Durum='Aktif' olan tum kayitlari dondur"""
        wb = openpyxl.load_workbook(EXCEL_PATH)
        aktifler = []
        for sname in wb.sheetnames:
            if sname == "Odalar":
                continue
            ws = wb[sname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[COL_DURUM - 1] == "Aktif":
                    aktifler.append({
                        "id": row[COL_ID - 1],
                        "tc": row[COL_TC - 1],
                        "isim": row[COL_ISIM - 1],
                        "soyisim": row[COL_SOYISIM - 1],
                        "oda": row[COL_ODA - 1],
                        "giris": row[COL_GIRIS - 1],
                        "cikis": row[COL_CIKIS - 1],
                        "odeme_yon": row[COL_ODEME_YON - 1],
                        "odeme_tut": row[COL_ODEME_TUT - 1],
                        "durum": row[COL_DURUM - 1],
                        "sheet": sname
                    })
        return aktifler

    def cikis_yaptir(self, kayit_id, sheet_name):
        """Musteriye cikis yaptir"""
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        bugun = datetime.now().strftime("%d/%m/%Y")

        for row in ws.iter_rows(min_row=2):
            if str(row[COL_ID - 1].value) == str(kayit_id):
                oda = row[COL_ODA - 1].value
                row[COL_CIKIS - 1].value = bugun
                row[COL_DURUM - 1].value = "Cikis Yapti"
                self._update_oda_durumu(wb, oda, "Musait")
                wb.save(EXCEL_PATH)
                return True
        return False

    def kayit_guncelle(self, kayit_id, sheet_name, data):
        """Mevcut kaydi guncelle"""
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2):
            if str(row[COL_ID - 1].value) == str(kayit_id):
                eski_oda = row[COL_ODA - 1].value
                yeni_oda = data.get("oda", eski_oda)

                row[COL_TC - 1].value = data.get("tc", row[COL_TC - 1].value)
                row[COL_ISIM - 1].value = data.get("isim", row[COL_ISIM - 1].value)
                row[COL_SOYISIM - 1].value = data.get("soyisim", row[COL_SOYISIM - 1].value)
                row[COL_ODA - 1].value = yeni_oda
                row[COL_GIRIS - 1].value = data.get("giris", row[COL_GIRIS - 1].value)
                row[COL_CIKIS - 1].value = data.get("cikis", row[COL_CIKIS - 1].value)
                row[COL_ODEME_YON - 1].value = data.get("odeme_yon", row[COL_ODEME_YON - 1].value)
                row[COL_ODEME_TUT - 1].value = data.get("odeme_tut", row[COL_ODEME_TUT - 1].value)

                # Oda degistiyse durumlari guncelle
                if str(eski_oda) != str(yeni_oda):
                    self._update_oda_durumu(wb, eski_oda, "Musait")
                    self._update_oda_durumu(wb, yeni_oda, "Dolu")

                wb.save(EXCEL_PATH)
                return True
        return False

    def get_odalar(self):
        """Tum odalari dondur"""
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Odalar" not in wb.sheetnames:
            return []
        ws = wb["Odalar"]
        odalar = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                odalar.append({"no": str(row[0]), "durum": str(row[1]) if row[1] else "Musait"})
        return odalar

    def get_musait_odalar(self):
        return [o["no"] for o in self.get_odalar() if o["durum"] == "Musait"]

    def oda_ekle(self, oda_no):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Odalar" not in wb.sheetnames:
            ws = wb.create_sheet("Odalar")
            ws["A1"] = "Oda Numarasi"
            ws["B1"] = "Durum"
        ws = wb["Odalar"]
        # Zaten varsa ekleme
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(oda_no):
                return False
        ws.append([str(oda_no), "Musait"])
        wb.save(EXCEL_PATH)
        return True

    def oda_sil(self, oda_no):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Odalar" not in wb.sheetnames:
            return False
        ws = wb["Odalar"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(oda_no):
                ws.delete_rows(i)
                wb.save(EXCEL_PATH)
                return True
        return False

    def get_tum_kayitlar(self, filtre=None):
        """Tum gecmis kayitlari dondur, filtrelenebilir"""
        wb = openpyxl.load_workbook(EXCEL_PATH)
        kayitlar = []
        for sname in wb.sheetnames:
            if sname == "Odalar":
                continue
            # Yil ve donem ayristir
            parts = sname.split("_")
            yil = parts[0] if len(parts) > 0 else ""
            donem = parts[1] if len(parts) > 1 else ""

            ws = wb[sname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                kayit = {
                    "id": row[COL_ID - 1],
                    "tc": row[COL_TC - 1],
                    "isim": row[COL_ISIM - 1],
                    "soyisim": row[COL_SOYISIM - 1],
                    "oda": row[COL_ODA - 1],
                    "giris": row[COL_GIRIS - 1],
                    "cikis": row[COL_CIKIS - 1],
                    "odeme_yon": row[COL_ODEME_YON - 1],
                    "odeme_tut": row[COL_ODEME_TUT - 1],
                    "durum": row[COL_DURUM - 1],
                    "sheet": sname,
                    "yil": yil,
                    "donem": donem
                }

                if filtre:
                    arama = filtre.get("arama", "").lower()
                    yil_f = filtre.get("yil", "Tumu")
                    donem_f = filtre.get("donem", "Tumu")
                    durum_f = filtre.get("durum", "Tumu")

                    if arama:
                        isim_soy = f"{kayit['isim']} {kayit['soyisim']}".lower()
                        if (arama not in isim_soy and
                                arama not in str(kayit['tc']).lower() and
                                arama not in str(kayit['id']).lower()):
                            continue

                    if yil_f != "Tumu" and kayit["yil"] != yil_f:
                        continue
                    if donem_f != "Tumu" and kayit["donem"] != donem_f:
                        continue
                    if durum_f != "Tumu" and kayit["durum"] != durum_f:
                        continue

                kayitlar.append(kayit)
        return kayitlar

    def get_autocomplete_data(self):
        """Autocomplete icin tum isim/soyisim/TC listesi"""
        kayitlar = self.get_tum_kayitlar()
        kisiler = {}
        for k in kayitlar:
            key = str(k.get("tc", ""))
            if key and key not in kisiler:
                kisiler[key] = {
                    "tc": k.get("tc", ""),
                    "isim": k.get("isim", ""),
                    "soyisim": k.get("soyisim", ""),
                    "oda": k.get("oda", "")
                }
            # En son odayi guncelle
            elif key:
                kisiler[key]["oda"] = k.get("oda", "")
        return list(kisiler.values())

    def get_yillar(self):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        yillar = set()
        for sname in wb.sheetnames:
            if sname != "Odalar" and "_" in sname:
                yillar.add(sname.split("_")[0])
        return sorted(list(yillar), reverse=True)

    def check_backup_needed(self):
        config = self._load_config()
        son_yedek = config.get("son_yedekleme")
        if not son_yedek:
            return True
        try:
            son_tarih = datetime.strptime(son_yedek, "%Y-%m-%d").date()
            fark = (date.today() - son_tarih).days
            return fark > 30
        except:
            return True

    def update_backup_date(self):
        config = self._load_config()
        config["son_yedekleme"] = date.today().strftime("%Y-%m-%d")
        self._save_config(config)

    def get_backup_date(self):
        config = self._load_config()
        return config.get("son_yedekleme", "Hic yedeklenmedi")
